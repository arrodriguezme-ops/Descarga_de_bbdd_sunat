"""
ihh_excel_export.py

Exporta los resultados de ihh_concentracion.py a un Excel formateado (estilo
tomado del script original en R: encabezados azul oscuro con texto blanco,
participaciones en formato porcentaje, secciones tituladas).
"""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

AZUL = "FF002060"
BLANCO = "FFFFFFFF"

FUENTE_HEADER = Font(name="Arial", size=10, bold=True, color=BLANCO)
FUENTE_SECCION = Font(name="Arial", size=12, bold=True, color=AZUL)
FUENTE_TITULO = Font(name="Arial", size=11, bold=True, color=AZUL)
FUENTE_TEXTO = Font(name="Arial", size=10)
RELLENO_HEADER = PatternFill("solid", fgColor=AZUL)
BORDE_BLANCO = Border(*[Side(style="thin", color=BLANCO)] * 4)


def _escribir_df(ws, df: pd.DataFrame, fila_inicio: int, con_header: bool = True) -> int:
    """Escribe un DataFrame (con su indice como primera columna) a partir de
    fila_inicio. Devuelve la fila siguiente libre."""
    columnas = [df.index.name or ""] + [str(c) for c in df.columns]

    fila = fila_inicio
    if con_header:
        for j, nombre in enumerate(columnas, start=1):
            celda = ws.cell(row=fila, column=j, value=nombre)
            celda.font = FUENTE_HEADER
            celda.fill = RELLENO_HEADER
            celda.alignment = Alignment(horizontal="center")
            celda.border = BORDE_BLANCO
        fila += 1

    for idx, fila_datos in df.iterrows():
        ws.cell(row=fila, column=1, value=str(idx)).font = FUENTE_TEXTO
        for j, valor in enumerate(fila_datos, start=2):
            celda = ws.cell(row=fila, column=j)
            if pd.isna(valor):
                celda.value = None
            elif str(idx) == "IHH":
                celda.value = round(float(valor), 4)
                celda.number_format = "0.0000"
            elif "Total" in str(idx):
                celda.value = round(float(valor), 2)
                celda.number_format = "#,##0.00"
            else:
                celda.value = float(valor)
                celda.number_format = "0.00%"
            celda.font = FUENTE_TEXTO
        fila += 1

    return fila + 1


def _escribir_metricas(ws, metricas: pd.DataFrame, fila_inicio: int) -> int:
    fila = fila_inicio
    encabezados = ["Métrica", "Valor (fracción)", "Valor (x10,000)"]
    for j, nombre in enumerate(encabezados, start=1):
        celda = ws.cell(row=fila, column=j, value=nombre)
        celda.font = FUENTE_HEADER
        celda.fill = RELLENO_HEADER
        celda.alignment = Alignment(horizontal="center")
        celda.border = BORDE_BLANCO
    fila += 1

    for _, r in metricas.iterrows():
        es_ms = str(r["Métrica"]).startswith("MS")
        ws.cell(row=fila, column=1, value=r["Métrica"]).font = FUENTE_TEXTO
        c2 = ws.cell(row=fila, column=2, value=round(float(r["Valor (fracción)"]), 4))
        c2.number_format = "0.00%" if es_ms else "0.0000"
        c2.font = FUENTE_TEXTO
        if pd.notna(r.get("Valor (x10,000)")):
            c3 = ws.cell(row=fila, column=3, value=round(float(r["Valor (x10,000)"]), 2))
            c3.number_format = "#,##0.00"
            c3.font = FUENTE_TEXTO
        fila += 1
    return fila + 1


def exportar_excel_ihh(resultados: list, ruta_salida: Path) -> Path:
    """resultados: lista de ihh_concentracion.ResultadoVista"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Resultados"
    ws.sheet_view.showGridLines = False

    fila = 1
    for resultado in resultados:
        titulo_seccion = f"{resultado.nombre_vista} — {resultado.nombre_variable}"
        celda = ws.cell(row=fila, column=1, value=titulo_seccion)
        celda.font = FUENTE_SECCION
        fila += 2

        celda = ws.cell(row=fila, column=1, value="Participación de mercado por grupo económico")
        celda.font = FUENTE_TITULO
        fila += 1
        fila = _escribir_df(ws, resultado.tabla_ancha, fila, con_header=True)

        celda = ws.cell(row=fila, column=1, value="IHH por año")
        celda.font = FUENTE_TITULO
        fila += 1
        ihh_tabla = resultado.ihh_por_anio.set_index("anio")[["ihh", "ihh_x10000", "interpretacion"]]
        ihh_tabla.columns = ["IHH (fracción)", "IHH (x10,000)", "Interpretación"]
        for j, nombre in enumerate(["Año"] + list(ihh_tabla.columns), start=1):
            c = ws.cell(row=fila, column=j, value=nombre)
            c.font = FUENTE_HEADER
            c.fill = RELLENO_HEADER
            c.border = BORDE_BLANCO
        fila += 1
        for anio, r in ihh_tabla.iterrows():
            ws.cell(row=fila, column=1, value=int(anio)).font = FUENTE_TEXTO
            c2 = ws.cell(row=fila, column=2, value=round(float(r["IHH (fracción)"]), 4))
            c2.number_format = "0.0000"
            c2.font = FUENTE_TEXTO
            c3 = ws.cell(row=fila, column=3, value=round(float(r["IHH (x10,000)"]), 2))
            c3.number_format = "#,##0.00"
            c3.font = FUENTE_TEXTO
            ws.cell(row=fila, column=4, value=r["Interpretación"]).font = FUENTE_TEXTO
            fila += 1
        fila += 2

        if resultado.metricas_fusion is not None:
            titulo = f"Análisis de fusión: {resultado.grupo_adquiriente} + {resultado.grupo_objetivo}"
            celda = ws.cell(row=fila, column=1, value=titulo)
            celda.font = FUENTE_TITULO
            fila += 1
            fila = _escribir_metricas(ws, resultado.metricas_fusion, fila)

        fila += 1  # espacio entre secciones

    for col in range(1, 8):
        ws.column_dimensions[get_column_letter(col)].width = 26

    ruta_salida = Path(ruta_salida)
    ruta_salida.parent.mkdir(parents=True, exist_ok=True)
    wb.save(ruta_salida)
    return ruta_salida
